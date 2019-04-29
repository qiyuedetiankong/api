#-*- coding:utf-8-*-
# @Time    :2019/4/15
# @Author  :武汉-七月
# @File    :excelparse.py
# @Software:PyCharm Community Edition
from openpyxl import load_workbook
import json
from homework.API_0423.common import configparse
class ExcelParse:
    """
    init 构造函数打开工作簿，工作表
    get_case 获取用例
    write_result 写入测试数据，回填测试结果

    """
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(file_name)
        self.sheet = self.wb[sheet_name]

    def get_case(self):
        max_row = self.sheet.max_row
        cases = []
        for row in range(2, max_row+1):
            data={}
            data["case_id"] = self.sheet.cell(row,1).value
            data["url"] = configparse.ReadConfig().get_cfg_value("api","pre_url")+self.sheet.cell(row, 3).value
            data["data"] = json.loads(self.sheet.cell(row, 4).value,encoding="utf-8")
            data["method"] = self.sheet.cell(row, 5).value
            data["expected"] = self.sheet.cell(row, 6).value
            data["sql"] = self.sheet.cell(row,9).value
            data["title"] = self.sheet.cell(row,2).value
            cases.append(data)

        self.wb.close()
        return cases

    def write_result(self, row, actual, result ):
        sheet=self.wb[self.sheet_name]
        sheet.cell(row,7,str(actual))# 将获取的响应值转化成str存入excel中,字典不能存在excel中
        sheet.cell(row,8,result)
        self.wb.save(self.file_name)
        self.wb.close()

if __name__ == '__main__':

  data2 = ExcelParse("../data/cases.xlsx","login").write_result(4,{"NihOn":"你好"},"pass")
  data = ExcelParse("../data/cases.xlsx", "add").get_case()
  for item in data:
      print(item)
  data = ExcelParse("../data/cases.xlsx", "invest").get_case()
  for item in data:
   print(item)

